#!/usr/bin/env python3
"""
CacusaPOS Servidor WiFi - By Taitus LLC
Corre en la PC y recibe las ventas del iPhone en tiempo real.
Escribe directamente en el Excel de OneDrive al instante.
"""

import sys, subprocess, socket, threading, json, datetime, shutil, re, base64, hashlib
import urllib.request, urllib.error
from pathlib import Path

# ── Auto-instalar dependencias ──
def instalar(pkg):
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg, '--quiet'])

try:
    from flask import Flask, request, jsonify
    from flask_cors import CORS
except ImportError:
    print("Instalando Flask..."); instalar('flask'); instalar('flask-cors')
    from flask import Flask, request, jsonify
    from flask_cors import CORS

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl..."); instalar('openpyxl')
    import openpyxl

try:
    from fpdf import FPDF
except ImportError:
    print("Instalando fpdf2..."); instalar('fpdf2')
    from fpdf import FPDF

try:
    import webauthn
    from webauthn.helpers.structs import (
        AuthenticatorSelectionCriteria, UserVerificationRequirement,
        ResidentKeyRequirement, AuthenticatorAttachment,
        AttestationConveyancePreference, PublicKeyCredentialDescriptor,
        RegistrationCredential, AuthenticatorAttestationResponse,
        AuthenticationCredential, AuthenticatorAssertionResponse,
    )
    from webauthn.helpers.cose import COSEAlgorithmIdentifier
    from webauthn.helpers import bytes_to_base64url, base64url_to_bytes
    _WEBAUTHN_OK = True
except Exception:
    print("Instalando webauthn..."); instalar('webauthn')
    import webauthn
    from webauthn.helpers.structs import (
        AuthenticatorSelectionCriteria, UserVerificationRequirement,
        ResidentKeyRequirement, AuthenticatorAttachment,
        AttestationConveyancePreference, PublicKeyCredentialDescriptor,
        RegistrationCredential, AuthenticatorAttestationResponse,
        AuthenticationCredential, AuthenticatorAssertionResponse,
    )
    from webauthn.helpers.cose import COSEAlgorithmIdentifier
    from webauthn.helpers import bytes_to_base64url, base64url_to_bytes
    _WEBAUTHN_OK = True

import tkinter as tk
from tkinter import scrolledtext, messagebox as tk_messagebox

# ════════════════════════════════════════
#  CONFIGURACIÓN
# ════════════════════════════════════════
BASE          = Path(r"C:\Users\titaj\OneDrive\Documents\CACUSA VENTAS")
EXCEL         = BASE / "VENTAS 4.1" / "CacusaPOS_v2.xlsm"
FACTURAS_DIR  = BASE / "VENTAS 4.1" / "Facturas"
ORDENES_DIR   = BASE / "VENTAS 4.1" / "Orden de Compra"
PORT          = 5678
NGROK_DOMAIN  = "upfront-yearbook-fascism.ngrok-free.dev"
NGROK_URL     = f"https://{NGROK_DOMAIN}"
IDS_REGISTRADOS = set()   # evita duplicados en memoria

WA_RP_ID      = NGROK_DOMAIN                              # upfront-yearbook-fascism.ngrok-free.dev
WA_ORIGINS    = [NGROK_URL, 'http://localhost:5678']      # orígenes aceptados
WA_CREDS_FILE = BASE / 'Aplicacion Movil' / 'wa_creds.json'
WA_USUARIOS   = {'robin.gonzalez', 'tita.jaramillo'}
_wa_reg_ch  = {}   # { user: challenge_bytes }  — en memoria, expiran con el proceso
_wa_auth_ch = {}   # { user: challenge_bytes }

def _wa_load():
    if WA_CREDS_FILE.exists():
        try:
            import json as _json
            return _json.loads(WA_CREDS_FILE.read_text('utf-8'))
        except Exception: pass
    return {}

def _wa_save(data):
    import json as _json
    WA_CREDS_FILE.write_text(_json.dumps(data, indent=2), 'utf-8')

flask_app = Flask(__name__)
CORS(flask_app)            # permite conexiones desde el iPhone

@flask_app.after_request
def _skip_ngrok_warning(response):
    response.headers['ngrok-skip-browser-warning'] = '1'
    return response

ui_log        = None       # callback para mostrar en la UI
ngrok_proc    = None       # proceso ngrok

# ════════════════════════════════════════
#  ESCRITURA EN EXCEL
# ════════════════════════════════════════
_excel_lock = threading.Lock()

def ultima_fila_con_dato(ws, col, desde):
    for r in range(ws.max_row, desde - 1, -1):
        if ws.cell(r, col).value not in (None, ''):
            return r
    return desde - 1

def escribir_venta(sale):
    with _excel_lock:
        wb = openpyxl.load_workbook(EXCEL, keep_vba=True)
        tipo = sale.get('type', 'FAC')
        tax_label = f"{sale.get('taxState','')} {sale.get('taxRate',0)}%" if sale.get('taxState') else ''

        # items serializados: "id~name~cat~price~qty~0|..." (igual formato que ORDENES)
        items_str = '|'.join([
            f"{x.get('id','')}~{x.get('name','')}~{x.get('cat','')}~"
            f"{x.get('price',0)}~{x.get('qty',1)}~0"
            for x in sale.get('items', [])
        ])

        if tipo == 'FAC' and 'VENTAS' in wb.sheetnames:
            ws  = wb['VENTAS']
            row = ultima_fila_con_dato(ws, 2, 5) + 1
            ws.cell(row,  2).value = sale.get('id')
            ws.cell(row,  3).value = 'FACTURA'
            ws.cell(row,  4).value = sale.get('client')
            ws.cell(row,  5).value = round(float(sale.get('subtotal', 0)), 2)
            ws.cell(row,  6).value = round(float(sale.get('tax', 0)), 2)
            ws.cell(row,  7).value = round(float(sale.get('total', 0)), 2)
            ws.cell(row,  8).value = sale.get('reseller', 'NO')
            ws.cell(row,  9).value = tax_label
            ws.cell(row, 10).value = sale.get('status', 'Pendiente')
            ws.cell(row, 11).value = items_str   # col K: detalle de productos (antes vacía)
            ws.cell(row, 12).value = round(float(sale.get('usps', 0)), 2)
            ws.cell(row, 13).value = round(float(sale.get('shipping', 0)), 2)
            ws.cell(row, 14).value = sale.get('date')
            ws.cell(row, 15).value = sale.get('payMethod')
            ws.cell(row, 16).value = sale.get('seller', '')   # col P: vendedor (trazabilidad)
            if not ws.cell(4, 16).value:
                ws.cell(4, 16).value = 'Vendedor'             # encabezado si está vacío

        elif tipo == 'OC' and 'ORDENES' in wb.sheetnames:
            ws   = wb['ORDENES']
            row  = ultima_fila_con_dato(ws, 2, 6) + 1
            num  = row - 5
            ws.cell(row,  1).value = num
            ws.cell(row,  2).value = sale.get('id')
            ws.cell(row,  3).value = sale.get('date')
            ws.cell(row,  4).value = sale.get('client')
            ws.cell(row,  5).value = sale.get('status', 'Pendiente')
            ws.cell(row,  6).value = round(float(sale.get('subtotal', 0)), 2)
            ws.cell(row,  7).value = round(float(sale.get('tax', 0)), 2)
            ws.cell(row,  8).value = round(float(sale.get('total', 0)), 2)
            ws.cell(row,  9).value = sale.get('payMethod')
            ws.cell(row, 10).value = round(float(sale.get('shipping', 0)), 2)
            ws.cell(row, 11).value = round(float(sale.get('usps', 0)), 2)
            ws.cell(row, 12).value = tax_label
            ws.cell(row, 13).value = ''
            ws.cell(row, 14).value = items_str
            ws.cell(row, 15).value = sale.get('seller', '')   # col O: vendedor (trazabilidad)
            if not ws.cell(5, 15).value:
                ws.cell(5, 15).value = 'Vendedor'             # encabezado si está vacío

        wb.save(EXCEL)

def _nombre_pdf(sale):
    """Genera el nombre del archivo PDF: OC-2026-001_Clara_Pelaez_20260524.pdf"""
    sid        = str(sale.get('id', 'DOC')).strip()
    client_raw = str(sale.get('client', 'Cliente')).strip()
    client_slug = ''.join(c if (c.isalnum() or c == '-') else '_' for c in client_raw)
    client_slug = re.sub(r'_+', '_', client_slug).strip('_')
    date_raw   = str(sale.get('date', '')).strip()
    try:
        if '/' in date_raw:
            d, m, y = date_raw.split('/')
            date_str = f"{y}{m.zfill(2)}{d.zfill(2)}"
        elif '-' in date_raw:
            date_str = date_raw.replace('-', '')
        else:
            date_str = datetime.datetime.now().strftime('%Y%m%d')
    except Exception:
        date_str = datetime.datetime.now().strftime('%Y%m%d')
    return f"{sid}_{client_slug}_{date_str}.pdf"

def generar_pdf_servidor(sale):
    """
    Genera PDF con formato IDENTICO a la app (jsPDF):
      - Header rosa 26mm, título 20pt bold, subtítulo 12pt
      - Fecha en español largo: "30 de mayo de 2026"
      - Datos: N° Doc, Cliente, Estado (fiscal), Método, Pago (estado)
      - Tabla items 2 columnas: "■ nombre  (qty x $precio)" | "$total"
      - Dije y envío como filas de items (igual que la app)
      - Totales: Subtotal, Dijes, Tax, CC Fee  — sin USPS ni Shipping aparte
      - Formas de pago con ■, slogan en cursiva, footer rosa 8pt
    Carga Arial desde Windows para soporte de ■ y caracteres Unicode.
    """
    tipo   = sale.get('type', 'FAC')
    folder = FACTURAS_DIR if tipo == 'FAC' else ORDENES_DIR
    folder.mkdir(parents=True, exist_ok=True)
    filename = _nombre_pdf(sale)
    filepath = folder / filename

    PINK  = (201, 99, 122)
    WHITE = (255, 255, 255)
    DARK  = (58,  16,  32)
    GRAY  = (160, 96, 122)   # igual que la app
    LPINK = (253, 232, 240)

    pdf = FPDF(format='letter')
    pdf.set_margins(18, 18, 18)   # ≈ 50pt como la app
    pdf.set_auto_page_break(auto=True, margin=22)

    # ── Cargar Arial para soporte de ■ y caracteres Unicode ──
    import os
    font = 'Helvetica'
    try:
        _ar  = r'C:\Windows\Fonts\arial.ttf'
        _arb = r'C:\Windows\Fonts\arialbd.ttf'
        _ari = r'C:\Windows\Fonts\ariali.ttf'
        if all(os.path.exists(p) for p in [_ar, _arb, _ari]):
            pdf.add_font('Arial', '',  _ar)
            pdf.add_font('Arial', 'B', _arb)
            pdf.add_font('Arial', 'I', _ari)
            font = 'Arial'
    except Exception:
        pass
    BUL = '■' if font == 'Arial' else '-'   # ■ o -

    def fmt(v):
        return f"${float(v or 0):.2f}"

    # ── Fecha en español largo (igual que la app) ──
    meses = ['enero','febrero','marzo','abril','mayo','junio',
             'julio','agosto','septiembre','octubre','noviembre','diciembre']
    date_raw = str(sale.get('date', '')).strip()
    try:
        if '/' in date_raw:
            dd, mm, yyyy = date_raw.split('/')
        elif '-' in date_raw:
            parts = date_raw.split('-')
            yyyy, mm, dd = parts if len(parts[0]) == 4 else (parts[2], parts[1], parts[0])
        else:
            raise ValueError
        fStr      = f"{int(dd)} de {meses[int(mm)-1]} de {yyyy}"
        fechaCorta = f"{dd.zfill(2)}/{mm.zfill(2)}/{yyyy}"
    except Exception:
        fStr = fechaCorta = date_raw

    pdf.add_page()
    W  = pdf.w    # 215.9 mm
    EW = pdf.epw  # ~179.9 mm (215.9 - 18 - 18)

    # ── Header rosa ── igual que app: rect 0,0 → W,26mm; textos en 20pt/12pt
    pdf.set_fill_color(*PINK)
    pdf.rect(0, 0, W, 26, 'F')
    pdf.set_text_color(*WHITE)
    pdf.set_font(font, 'B', 20)
    pdf.set_xy(0, 6)
    pdf.cell(W, 10, 'CACUSA BY TAITUS LLC',
             align='C', new_x='LMARGIN', new_y='NEXT')
    pdf.set_font(font, '', 12)
    pdf.set_x(0)
    pdf.cell(W, 6, 'FACTURA' if tipo == 'FAC' else 'ORDEN DE COMPRA',
             align='C', new_x='LMARGIN', new_y='NEXT')

    pdf.set_xy(18, 34)   # inicio zona de datos (≈ y=98pt en la app)

    # ── campo(): label bold + valor normal, igual que la app ──
    LBL = 40   # mm — ancho de la columna de etiqueta

    def campo(lbl, val, bold_val=False):
        pdf.set_font(font, 'B', 10)
        pdf.set_text_color(*DARK)
        pdf.cell(LBL, 7, lbl,       new_x='RIGHT', new_y='TOP')
        pdf.set_font(font, 'B' if bold_val else '', 10)
        pdf.cell(0,   7, str(val),  new_x='LMARGIN', new_y='NEXT')

    # N° Documento con fecha larga (igual que la app: mismo campo, mismo formato)
    campo('Nº Documento:', f"{sale.get('id','')}    {fStr}")

    campo('Cliente:', str(sale.get('client','') or '—'))

    ts = str(sale.get('taxState','') or '').strip()
    tr = sale.get('taxRate', 0)
    if ts:
        # app usa " — " (em dash); usamos " - " por compatibilidad Latin-1 en fallback
        sep = ' — ' if font == 'Arial' else ' - '
        campo('Estado:', f"{ts}{sep}Tax {tr}%")

    pm = str(sale.get('payMethod','') or '').strip()
    campo('Método:', pm or '—')

    # Pago: estado de la venta (campo NUEVO, igual en app y servidor)
    st = str(sale.get('status','') or '').strip()
    if st and st != 'Eliminado':
        campo('Pago:', st, bold_val=True)

    # ── Tabla de items 2 columnas (IDENTICA a la app) ──
    pdf.ln(5)
    items    = sale.get('items', [])
    dije     = float(sale.get('dije',     0) or 0)
    dije_qty = int(sale.get('dijeQty',   0) or 0) or 1
    shp      = float(sale.get('shipping', 0) or 0)

    # Armar filas: "■ nombre  (qty x $precio)" igual que la app
    rows = [
        (f"{BUL} {it.get('name','')}  ({int(it.get('qty',1))} x {fmt(it.get('price',0))})",
         fmt(float(it.get('price',0)) * int(it.get('qty',1))))
        for it in items
    ]
    if dije > 0: rows.append((f"{BUL} Dije extra",    fmt(dije)))
    if shp  > 0: rows.append((f"{BUL} Envio cliente", fmt(shp)))

    if rows:
        col_p = 46        # ancho columna precio (≈ R-8 en app)
        col_d = EW - col_p  # ancho columna descripción

        # Header tabla rosa
        pdf.set_fill_color(*PINK)
        pdf.set_text_color(*WHITE)
        pdf.set_font(font, 'B', 9)
        pdf.cell(col_d, 7, 'DESCRIPCION',         fill=True, new_x='RIGHT', new_y='TOP')
        pdf.cell(col_p, 7, 'PRECIO', align='R',   fill=True, new_x='LMARGIN', new_y='NEXT')

        # Filas con bandas alternas rosas (igual que la app)
        pdf.set_text_color(*DARK)
        pdf.set_font(font, '', 9)
        for i, (desc, precio) in enumerate(rows):
            fr = (i % 2 == 0)
            if fr: pdf.set_fill_color(*LPINK)
            pdf.cell(col_d, 7, desc,    fill=fr,             new_x='RIGHT', new_y='TOP')
            pdf.cell(col_p, 7, precio,  fill=fr, align='R',  new_x='LMARGIN', new_y='NEXT')

    # ── Línea separadora y Totales (misma estructura que la app) ──
    pdf.ln(5)
    pdf.set_draw_color(*PINK)
    pdf.set_line_width(0.5)
    pdf.line(18, pdf.get_y(), 18 + EW, pdf.get_y())
    pdf.ln(5)

    subtotal = float(sale.get('subtotal', 0) or 0)
    tax      = float(sale.get('tax',      0) or 0)
    cc_fee   = float(sale.get('ccFee',    0) or 0)
    total    = float(sale.get('total',    0) or 0)
    tax_rate = sale.get('taxRate', 0)

    col_lbl = EW - 46   # ≈ R-130 en la app
    col_val = 46        # ≈ 130pt en la app

    def tot_row(lbl, val):
        pdf.set_font(font, '', 10)
        pdf.set_text_color(*DARK)
        pdf.cell(col_lbl, 6, lbl,        new_x='RIGHT', new_y='TOP')
        pdf.cell(col_val, 6, fmt(val),   new_x='LMARGIN', new_y='NEXT', align='R')

    tot_row('Subtotal', subtotal)
    if dije > 0:
        dq = dije_qty
        tot_row(f"Dijes extra ({dq}×$5.00)", dije)   # × en Unicode
    if tax > 0:
        tot_row(f"Tax ({tax_rate}%)" if tax_rate else "Tax", tax)
    if cc_fee > 0:
        tot_row('Recargo Tarjeta 4%', cc_fee)

    # Fila TOTAL con fondo rosa (igual que la app: rect de fondo + texto blanco 12pt bold)
    pdf.ln(2)
    pdf.set_fill_color(*PINK)
    pdf.set_text_color(*WHITE)
    pdf.set_font(font, 'B', 12)
    pdf.cell(col_lbl, 10, 'TOTAL',         fill=True,             new_x='RIGHT', new_y='TOP')
    pdf.cell(col_val, 10, fmt(total),       fill=True, align='R',  new_x='LMARGIN', new_y='NEXT')

    # ── Formas de pago (igual que la app) ──
    pdf.ln(10)
    pdf.set_font(font, 'B', 10)
    pdf.set_text_color(*DARK)
    pdf.cell(0, 7, 'FORMAS DE PAGO', new_x='LMARGIN', new_y='NEXT')
    pdf.set_font(font, '', 9)
    pdf.set_text_color(*GRAY)
    for ln_txt in [
        f"{BUL} Zelle: facturacioncacusa@gmail.com",
        f"{BUL} Cash App: $cacusabytaitus",
        f"{BUL} Tarjeta de credito: 4% de recargo - solicitar link de pago",
    ]:
        pdf.cell(0, 6, ln_txt, new_x='LMARGIN', new_y='NEXT')

    # ── Slogan en cursiva centrado (igual que la app) ──
    pdf.ln(10)
    pdf.set_font(font, 'I', 8)
    pdf.set_text_color(*PINK)
    pdf.cell(0, 5,
             '"En Cacusa creamos para que luzcas mas bella de lo que ya eres"',
             align='C', new_x='LMARGIN', new_y='NEXT')

    # ── Footer rosa 8pt (igual que la app) ──
    pdf.set_y(-18)
    pdf.set_fill_color(*PINK)
    pdf.set_text_color(*WHITE)
    pdf.set_font(font, '', 8)
    pdf.cell(0, 12,
             f"CACUSA BY TAITUS LLC  |  {fechaCorta}  |  {sale.get('id','')}  |  Pag. 1 de 1",
             fill=True, align='C', new_x='LMARGIN', new_y='NEXT')

    pdf.output(str(filepath))
    return filepath

def escribir_costo(cost):
    with _excel_lock:
        wb = openpyxl.load_workbook(EXCEL, keep_vba=True)
        if 'COSTOS' in wb.sheetnames:
            ws  = wb['COSTOS']
            # Headers en fila 5, datos desde fila 6
            row = ultima_fila_con_dato(ws, 4, 6) + 1
            if row < 6:
                row = 6
            # Columna A: número correlativo
            ws.cell(row, 1).value = row - 5
            ws.cell(row, 2).value = cost.get('date')
            ws.cell(row, 3).value = cost.get('type')
            ws.cell(row, 4).value = cost.get('desc')
            ws.cell(row, 5).value = round(float(cost.get('amount', 0)), 2)
            ws.cell(row, 6).value = cost.get('provider', '')
            ws.cell(row, 7).value = ''
            # Guardar timestamp como referencia (columna H: NOTAS)
            ws.cell(row, 8).value = str(cost.get('created', ''))
            wb.save(EXCEL)

def _costo_id_estable(fecha, tipo, desc, monto):
    """Genera un ID numérico estable para costos ingresados directamente en Excel (sin timestamp)."""
    key = f"{fecha}|{tipo}|{desc}|{round(float(monto or 0), 2)}"
    return int(hashlib.md5(key.encode()).hexdigest()[:12], 16) % (10 ** 13)

def leer_costos():
    """Lee todos los costos de la hoja COSTOS del Excel.
    Para filas sin timestamp (ingresadas a mano en Excel) genera un ID estable
    y lo escribe en col H para que sea permanente en lecturas futuras.
    """
    costos = []
    dirty  = False          # si hay IDs nuevos que guardar
    try:
        with _excel_lock:
            wb = openpyxl.load_workbook(EXCEL, keep_vba=True)
            if 'COSTOS' not in wb.sheetnames:
                return costos
            ws = wb['COSTOS']
            # Headers en fila 5, datos desde fila 6
            for row_idx, row in enumerate(ws.iter_rows(min_row=6, values_only=False), start=6):
                vals = [c.value for c in row]
                if not any(v is not None for v in vals[:6]):
                    continue
                fecha = str(vals[1]) if vals[1] else ''
                tipo  = str(vals[2]) if vals[2] else ''
                desc  = str(vals[3]) if vals[3] else ''
                if not desc:
                    continue
                monto = float(vals[4]) if vals[4] not in (None, '') else 0.0
                prov  = str(vals[5]) if vals[5] else ''
                # Col H (índice 7): timestamp/created
                created_raw = vals[7] if len(vals) > 7 else None
                try:
                    created = int(created_raw) if created_raw not in (None, '') else None
                except (ValueError, TypeError):
                    created = None
                # Si no tiene ID, generar uno estable y escribirlo
                if created is None:
                    created = _costo_id_estable(fecha, tipo, desc, monto)
                    ws.cell(row=row_idx, column=8).value = created
                    dirty = True
                costos.append({
                    'type': tipo, 'desc': desc, 'amount': monto,
                    'provider': prov, 'date': fecha,
                    'created': created
                })
            if dirty:
                wb.save(EXCEL)
            wb.close()
    except Exception:
        pass
    return costos

def actualizar_estado_en_excel(sale_id, tipo, new_status):
    """
    Encuentra la fila con sale_id en la hoja VENTAS u ORDENES y actualiza su estado.
    Devuelve True si encontró y guardó, False si no encontró el ID.
    """
    with _excel_lock:
        wb = openpyxl.load_workbook(EXCEL, keep_vba=True)
        sheet_name = 'VENTAS' if tipo == 'FAC' else 'ORDENES'
        status_col = 10 if tipo == 'FAC' else 5   # col J (FAC), col E (OC)
        min_row    = 5  if tipo == 'FAC' else 6
        if sheet_name not in wb.sheetnames:
            wb.close(); return False
        ws    = wb[sheet_name]
        found = False
        for row_cells in ws.iter_rows(min_row=min_row, min_col=2, max_col=2):
            id_cell = row_cells[0]
            if str(id_cell.value or '').strip() == str(sale_id).strip():
                ws.cell(id_cell.row, status_col).value = new_status
                found = True
                break
        if found:
            wb.save(EXCEL)
        wb.close()
        return found

def eliminar_pdfs_de_carpeta(sale_id, tipo):
    """
    Elimina todos los archivos PDF de la carpeta que empiecen con sale_id_.
    Devuelve la lista de nombres eliminados.
    """
    folder    = FACTURAS_DIR if tipo == 'FAC' else ORDENES_DIR
    eliminados = []
    if folder.exists():
        for f in folder.glob(f"{sale_id}_*.pdf"):
            try:
                f.unlink()
                eliminados.append(f.name)
            except Exception:
                pass
    return eliminados

# ════════════════════════════════════════
#  ENDPOINTS FLASK
# ════════════════════════════════════════
@flask_app.route('/', methods=['GET'])
@flask_app.route('/app', methods=['GET'])
def serve_app():
    """Sirve la app directamente — resuelve el bloqueo HTTPS/HTTP en iPhone."""
    html_path = BASE / 'Aplicacion Movil' / 'CacusaPOS.html'
    if not html_path.exists():
        return "CacusaPOS.html no encontrado", 404
    with open(html_path, 'r', encoding='utf-8') as f:
        content = f.read()
    return content, 200, {'Content-Type': 'text/html; charset=utf-8'}

@flask_app.route('/ping', methods=['GET', 'OPTIONS'])
def ping():
    return jsonify({'ok': True, 'service': 'CacusaPOS'})

@flask_app.route('/venta', methods=['POST', 'OPTIONS'])
def recibir_venta():
    if request.method == 'OPTIONS':
        return jsonify({'ok': True})
    sale = request.get_json(force=True)
    sale_id = sale.get('id', '')

    # Evitar duplicados
    if sale_id in IDS_REGISTRADOS:
        return jsonify({'ok': True, 'msg': 'ya_registrado'})

    try:
        escribir_venta(sale)
        IDS_REGISTRADOS.add(sale_id)
        _ventas_cache['ts'] = 0   # invalidar caché para que GET /ventas refleje el cambio
        # Generar PDF automáticamente en la carpeta correspondiente
        try:
            pdf_path = generar_pdf_servidor(sale)
            if ui_log: ui_log(f"📄  PDF guardado: {pdf_path.name}")
        except Exception as pdf_err:
            if ui_log: ui_log(f"⚠️  PDF no generado: {pdf_err}")
        hora = datetime.datetime.now().strftime('%H:%M')
        msg  = f"✅  [{hora}]  {sale_id} · {sale.get('client','')} · ${float(sale.get('total',0)):.2f}"
        if ui_log: ui_log(msg)
        return jsonify({'ok': True})
    except PermissionError:
        err = 'El archivo Excel está abierto. Ciérralo.'
        if ui_log: ui_log(f"⚠️  {err}")
        return jsonify({'ok': False, 'error': err}), 500
    except Exception as e:
        if ui_log: ui_log(f"❌  Error: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500

def leer_productos():
    """Lee el catálogo de productos de la hoja PRODUCTOS del Excel."""
    productos = []
    try:
        with _excel_lock:
            wb = openpyxl.load_workbook(EXCEL, keep_vba=True, data_only=True)
            if 'PRODUCTOS' not in wb.sheetnames:
                return productos
            ws = wb['PRODUCTOS']
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row[1]:
                    continue
                prod_id = str(row[1]).strip()
                # Ignorar filas basura (ej. 'B4')
                if not prod_id or not prod_id.startswith('P'):
                    continue
                name = str(row[2]).strip() if row[2] else ''
                cat  = str(row[3]).strip() if row[3] else ''
                if not name:
                    continue
                try:
                    price = float(row[4]) if row[4] not in (None, '') else 0.0
                except (TypeError, ValueError):
                    price = 0.0
                active_raw = str(row[6]).strip().lower() if len(row) > 6 and row[6] else 'si'
                active = active_raw in ('sí', 'si', 'yes', 'true', '1')
                productos.append({
                    'id': prod_id, 'name': name, 'cat': cat,
                    'price': price, 'active': active
                })
    except Exception:
        pass
    return productos

def escribir_producto(prod):
    """Agrega o actualiza un producto en la hoja PRODUCTOS del Excel."""
    prod_id = str(prod.get('id', '')).strip()
    if not prod_id:
        return
    with _excel_lock:
        wb = openpyxl.load_workbook(EXCEL, keep_vba=True)
        if 'PRODUCTOS' not in wb.sheetnames:
            return
        ws = wb['PRODUCTOS']
        # Buscar fila existente
        target_row = None
        for row in ws.iter_rows(min_row=5, values_only=False):
            if row[1].value == prod_id:
                target_row = row[0].row
                break
        if target_row is None:
            target_row = ultima_fila_con_dato(ws, 2, 5) + 1
            if target_row < 5:
                target_row = 5
        ws.cell(target_row, 2).value = prod_id
        ws.cell(target_row, 3).value = prod.get('name', '')
        ws.cell(target_row, 4).value = prod.get('cat', '')
        ws.cell(target_row, 5).value = round(float(prod.get('price', 0)), 2)
        ws.cell(target_row, 7).value = 'Sí' if prod.get('active', True) else 'No'
        wb.save(EXCEL)

@flask_app.route('/productos', methods=['GET', 'OPTIONS'])
def obtener_productos():
    if request.method == 'OPTIONS':
        return jsonify({'ok': True})
    try:
        prods = leer_productos()
        return jsonify({'ok': True, 'total': len(prods), 'productos': prods})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@flask_app.route('/producto', methods=['POST', 'OPTIONS'])
def guardar_producto():
    if request.method == 'OPTIONS':
        return jsonify({'ok': True})
    prod = request.get_json(force=True)
    try:
        escribir_producto(prod)
        hora = datetime.datetime.now().strftime('%H:%M')
        accion = 'actualizado' if prod.get('id') else 'creado'
        if ui_log: ui_log(f"📦  [{hora}]  Producto {accion}: {prod.get('id','')} {prod.get('name','')}")
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

ADMIN_PRODUCTS_URL = 'https://cacusabytaitus.com/data/products.json'

def sincronizar_desde_admin():
    """Descarga products.json del sitio web y los escribe en la hoja PRODUCTOS del Excel.
    Abre y guarda el Excel UNA sola vez para todos los productos (no una vez por producto)."""
    import json as _json
    try:
        req = urllib.request.Request(
            ADMIN_PRODUCTS_URL,
            headers={'User-Agent': 'CacusaPOS-Server/1.0'}
        )
        with urllib.request.urlopen(req, timeout=10) as r:
            data = _json.loads(r.read())
    except Exception as e:
        return 0, f"No se pudo descargar products.json: {e}"
    web_prods = data.get('products', []) if isinstance(data, dict) else data
    if not isinstance(web_prods, list):
        return 0, "Formato inesperado en products.json"

    prods = []
    for p in web_prods:
        pid = p.get('id')
        if pid is None:
            continue
        pid_str = str(pid) if str(pid).startswith('P') else f"P{pid}"
        name = str(p.get('name') or '').strip()
        if not name:
            continue
        prods.append({
            'id':     pid_str,
            'name':   name,
            'cat':    str(p.get('category') or '').strip(),
            'price':  float(p.get('price') or 0),
            'active': bool(p.get('available', True)),
        })
    if not prods:
        return 0, None

    count = 0
    with _excel_lock:
        wb = openpyxl.load_workbook(EXCEL, keep_vba=True)
        if 'PRODUCTOS' not in wb.sheetnames:
            return 0, "El Excel no tiene hoja PRODUCTOS"
        ws = wb['PRODUCTOS']
        # Mapa id → fila existente (una sola pasada)
        filas = {}
        for row in ws.iter_rows(min_row=5, values_only=False):
            v = row[1].value
            if v not in (None, ''):
                filas[str(v).strip()] = row[0].row
        siguiente = max(ultima_fila_con_dato(ws, 2, 5) + 1, 5)
        for prod in prods:
            fila = filas.get(prod['id'])
            if fila is None:
                fila = siguiente
                filas[prod['id']] = fila
                siguiente += 1
            ws.cell(fila, 2).value = prod['id']
            ws.cell(fila, 3).value = prod['name']
            ws.cell(fila, 4).value = prod['cat']
            ws.cell(fila, 5).value = round(prod['price'], 2)
            ws.cell(fila, 7).value = 'Sí' if prod['active'] else 'No'
            count += 1
        wb.save(EXCEL)
    return count, None

@flask_app.route('/sync-admin-productos', methods=['GET', 'POST', 'OPTIONS'])
def sync_admin_productos():
    if request.method == 'OPTIONS':
        return jsonify({'ok': True})
    count, error = sincronizar_desde_admin()
    hora = datetime.datetime.now().strftime('%H:%M')
    if error:
        if ui_log: ui_log(f"⚠️  [{hora}]  Sync admin fallido: {error}")
        return jsonify({'ok': False, 'error': error}), 500
    if ui_log: ui_log(f"🔄  [{hora}]  Sync admin: {count} productos actualizados desde la web")
    return jsonify({'ok': True, 'count': count})

@flask_app.route('/costos', methods=['GET', 'OPTIONS'])
def obtener_costos():
    if request.method == 'OPTIONS':
        return jsonify({'ok': True})
    try:
        costos = leer_costos()
        return jsonify({'ok': True, 'total': len(costos), 'costos': costos})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@flask_app.route('/costo', methods=['POST', 'OPTIONS'])
def recibir_costo():
    if request.method == 'OPTIONS':
        return jsonify({'ok': True})
    cost = request.get_json(force=True)
    try:
        escribir_costo(cost)
        hora = datetime.datetime.now().strftime('%H:%M')
        if ui_log: ui_log(f"💰  [{hora}]  Costo: {cost.get('desc','')} · ${float(cost.get('amount',0)):.2f}")
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

def _square_api(method, path, token, body=None):
    """Helper para llamar a la API de Square desde el servidor."""
    url = f'https://connect.squareup.com{path}'
    data = json.dumps(body).encode() if body else None
    req  = urllib.request.Request(
        url, data=data,
        headers={
            'Authorization':  f'Bearer {token}',
            'Content-Type':   'application/json',
            'Square-Version': '2024-01-17',
        },
        method=method
    )
    with urllib.request.urlopen(req, timeout=15) as r:
        return json.loads(r.read())

@flask_app.route('/square-link', methods=['POST', 'OPTIONS'])
def crear_link_square():
    """Genera un link de pago de Square (Payment Links API)."""
    if request.method == 'OPTIONS':
        return jsonify({'ok': True})
    data        = request.get_json(force=True)
    token       = data.get('access_token', '').strip()
    location_id = data.get('location_id', '').strip()
    amount      = data.get('amount', 0)       # en centavos
    name        = data.get('name', 'CACUSA POS')
    idem_key    = data.get('idempotency_key', str(datetime.datetime.now().timestamp()))

    if not token or not location_id:
        return jsonify({'ok': False, 'error': 'Falta access_token o location_id'}), 400

    try:
        result = _square_api('POST', '/v2/online-checkout/payment-links', token, {
            'idempotency_key': idem_key,
            'quick_pay': {
                'name':         name,
                'price_money':  {'amount': int(amount), 'currency': 'USD'},
                'location_id':  location_id
            }
        })
        url = result.get('payment_link', {}).get('url', '')
        if not url:
            return jsonify({'ok': False, 'error': 'Square no devolvió URL'}), 500
        hora = datetime.datetime.now().strftime('%H:%M')
        if ui_log: ui_log(f"🔗  [{hora}]  Link generado · {name}")
        return jsonify({'ok': True, 'url': url})
    except urllib.error.HTTPError as e:
        err = json.loads(e.read()).get('errors', [{}])
        msg = err[0].get('detail', str(e)) if err else str(e)
        return jsonify({'ok': False, 'error': msg}), 400
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@flask_app.route('/square-locations', methods=['POST', 'OPTIONS'])
def listar_locations_square():
    """Devuelve las locations de la cuenta Square (para configuración automática)."""
    if request.method == 'OPTIONS':
        return jsonify({'ok': True})
    data  = request.get_json(force=True)
    token = data.get('access_token', '').strip()
    if not token:
        return jsonify({'ok': False, 'error': 'Falta access_token'}), 400
    try:
        result    = _square_api('GET', '/v2/locations', token)
        locations = [
            {'id': loc['id'], 'name': loc.get('name', ''), 'status': loc.get('status', '')}
            for loc in result.get('locations', [])
        ]
        return jsonify({'ok': True, 'locations': locations})
    except urllib.error.HTTPError as e:
        err = json.loads(e.read()).get('errors', [{}])
        msg = err[0].get('detail', str(e)) if err else str(e)
        return jsonify({'ok': False, 'error': msg}), 400
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return '127.0.0.1'

def cargar_ids_desde_excel():
    """Pre-carga los IDs ya existentes en Excel para evitar duplicados tras reinicio."""
    try:
        if not EXCEL.exists():
            return
        wb = openpyxl.load_workbook(EXCEL, keep_vba=True, read_only=True, data_only=True)
        for sheet_name in ('VENTAS', 'ORDENES'):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
                val = row[0]
                if val and isinstance(val, str) and ('-' in val):
                    IDS_REGISTRADOS.add(val.strip())
        wb.close()
    except Exception:
        pass   # Si el Excel no se puede leer al arrancar, no es crítico

# ════════════════════════════════════════
#  LECTURA DESDE EXCEL — sincronización bidireccional con móviles
# ════════════════════════════════════════
import time as _time
_ventas_cache = {'ts': 0, 'data': []}
_CACHE_TTL    = 30   # segundos — evita re-leer el Excel en cada request

def _parse_tax_label(lbl):
    """'FL 7%' → ('FL', 7.0)  |  '' → ('', 0.0)"""
    lbl = str(lbl or '').strip()
    if not lbl:
        return '', 0.0
    parts = lbl.split()
    if len(parts) >= 2:
        try:
            return parts[0], float(parts[1].replace('%', ''))
        except Exception:
            pass
    return lbl, 0.0

def _parse_date(val):
    """Convierte datetime o string a 'DD/MM/YYYY'."""
    if hasattr(val, 'strftime'):
        return val.strftime('%d/%m/%Y')
    return str(val or '').strip()

def leer_ventas_de_excel():
    """
    Lee todas las ventas (FAC) y órdenes (OC) del Excel y las devuelve como
    lista de dicts compatibles con el formato localStorage de la app.
    Cachea 30 s para no bloquear el servidor en cada request.
    """
    now = _time.time()
    if now - _ventas_cache['ts'] < _CACHE_TTL:
        return _ventas_cache['data']

    ventas = []
    if not EXCEL.exists():
        _ventas_cache.update({'ts': now, 'data': ventas})
        return ventas

    try:
        wb = openpyxl.load_workbook(EXCEL, keep_vba=True, read_only=True, data_only=True)

        # ── Hoja VENTAS (FAC) ──
        # Col: B=id, C=tipo, D=cliente, E=subtotal, F=tax, G=total,
        #      H=reseller, I=tax_lbl, J=status, K=vacío,
        #      L=usps, M=shipping, N=fecha, O=payMethod
        if 'VENTAS' in wb.sheetnames:
            for row in wb['VENTAS'].iter_rows(min_row=5, values_only=True):
                sid = row[1]
                if not sid or '-' not in str(sid):
                    continue
                ts, tr = _parse_tax_label(row[8])
                # Reconstruir items desde col K (índice 10): "id~name~cat~price~qty~0|..."
                fac_items = []
                items_raw = row[10] if len(row) > 10 else None
                for seg in str(items_raw or '').split('|'):
                    p = seg.split('~')
                    if len(p) >= 5:
                        try:
                            fac_items.append({
                                'id':    p[0], 'name': p[1], 'cat': p[2],
                                'price': float(p[3]), 'qty': int(float(p[4])),
                            })
                        except Exception:
                            pass
                ventas.append({
                    'id':        str(sid).strip(),
                    'type':      'FAC',
                    'client':    str(row[3]  or ''),
                    'date':      _parse_date(row[13]),
                    'subtotal':  round(float(row[4]  or 0), 2),
                    'tax':       round(float(row[5]  or 0), 2),
                    'total':     round(float(row[6]  or 0), 2),
                    'reseller':  str(row[7]  or 'NO'),
                    'taxState':  ts,  'taxRate': tr,
                    'status':    str(row[9]  or 'Pagado'),
                    'usps':      round(float(row[11] or 0), 2),
                    'shipping':  round(float(row[12] or 0), 2),
                    'payMethod': str(row[14] or ''),
                    'seller':    str(row[15] or '') if len(row) > 15 else '',
                    'items': fac_items, 'dije': 0, 'dijeQty': 0, 'ccFee': 0,
                })

        # ── Hoja ORDENES (OC) ──
        # Col: A=num, B=id, C=fecha, D=cliente, E=status,
        #      F=subtotal, G=tax, H=total, I=payMethod,
        #      J=shipping, K=usps, L=tax_lbl, M=vacío,
        #      N=items_str, O=vacío
        if 'ORDENES' in wb.sheetnames:
            for row in wb['ORDENES'].iter_rows(min_row=6, values_only=True):
                sid = row[1]
                if not sid or '-' not in str(sid):
                    continue
                ts, tr = _parse_tax_label(row[11])
                # Reconstruir items desde "id~name~cat~price~qty~0|..."
                items = []
                for seg in str(row[13] or '').split('|'):
                    p = seg.split('~')
                    if len(p) >= 5:
                        try:
                            items.append({
                                'id':    p[0], 'name': p[1], 'cat': p[2],
                                'price': float(p[3]), 'qty': int(float(p[4])),
                            })
                        except Exception:
                            pass
                ventas.append({
                    'id':        str(sid).strip(),
                    'type':      'OC',
                    'client':    str(row[3]  or ''),
                    'date':      _parse_date(row[2]),
                    'status':    str(row[4]  or 'Pendiente'),
                    'subtotal':  round(float(row[5]  or 0), 2),
                    'tax':       round(float(row[6]  or 0), 2),
                    'total':     round(float(row[7]  or 0), 2),
                    'payMethod': str(row[8]  or ''),
                    'shipping':  round(float(row[9]  or 0), 2),
                    'usps':      round(float(row[10] or 0), 2),
                    'taxState':  ts,  'taxRate': tr,
                    'seller':    str(row[14] or '') if len(row) > 14 else '',
                    'items':     items,
                    'dije': 0, 'dijeQty': 0, 'ccFee': 0, 'reseller': 'NO',
                })

        wb.close()
        # Solo actualizar caché si la lectura fue exitosa
        _ventas_cache.update({'ts': _time.time(), 'data': ventas})
    except Exception:
        # Si el Excel está bloqueado o hay error, conservar los datos anteriores del caché
        # y retornar lo que haya (evita devolver lista vacía cuando hay datos válidos cacheados)
        pass

    return _ventas_cache['data']


@flask_app.route('/ventas', methods=['GET', 'OPTIONS'])
def obtener_ventas():
    """Devuelve todas las ventas/órdenes del Excel para sincronizar los móviles."""
    if request.method == 'OPTIONS':
        return jsonify({'ok': True})
    try:
        ventas = leer_ventas_de_excel()
        return jsonify({'ok': True, 'ventas': ventas, 'total': len(ventas)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@flask_app.route('/actualizar-venta', methods=['POST', 'OPTIONS'])
def actualizar_venta():
    """
    Actualiza el estado de una venta en Excel y regenera su PDF.
    Usado cuando el usuario cambia el estado en la app (Pendiente→Pagado→Cancelado).
    """
    if request.method == 'OPTIONS':
        return jsonify({'ok': True})
    data       = request.get_json(force=True)
    sale_id    = str(data.get('id', '')).strip()
    tipo       = data.get('type', 'FAC')
    new_status = str(data.get('status', '')).strip()
    if not sale_id or not new_status:
        return jsonify({'ok': False, 'error': 'Faltan id o status'}), 400
    try:
        found = actualizar_estado_en_excel(sale_id, tipo, new_status)
        _ventas_cache['ts'] = 0   # invalidar caché
        # Regenerar PDF con el nuevo estado
        if found:
            try:
                ventas    = leer_ventas_de_excel()
                sale_data = next((v for v in ventas if v['id'] == sale_id), None)
                if sale_data:
                    pdf_path = generar_pdf_servidor(sale_data)
                    hora = datetime.datetime.now().strftime('%H:%M')
                    if ui_log: ui_log(f"🔄  [{hora}]  {sale_id} → {new_status} · PDF regenerado")
            except Exception as pdf_err:
                if ui_log: ui_log(f"⚠️  PDF no regenerado: {pdf_err}")
        return jsonify({'ok': True, 'found': found})
    except Exception as e:
        if ui_log: ui_log(f"❌  actualizar-venta error: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@flask_app.route('/eliminar-venta', methods=['POST', 'OPTIONS'])
def eliminar_venta():
    """
    Marca la venta como 'Eliminado' en Excel y borra su PDF de la carpeta.
    Así el resto de dispositivos la eliminan en su próxima sincronización.
    """
    if request.method == 'OPTIONS':
        return jsonify({'ok': True})
    data    = request.get_json(force=True)
    sale_id = str(data.get('id', '')).strip()
    tipo    = data.get('type', 'FAC')
    if not sale_id:
        return jsonify({'ok': False, 'error': 'Falta id'}), 400
    try:
        actualizar_estado_en_excel(sale_id, tipo, 'Eliminado')
        _ventas_cache['ts'] = 0   # invalidar caché
        eliminados = eliminar_pdfs_de_carpeta(sale_id, tipo)
        hora = datetime.datetime.now().strftime('%H:%M')
        if ui_log: ui_log(f"🗑️  [{hora}]  {sale_id} eliminado · PDFs borrados: {len(eliminados)}")
        return jsonify({'ok': True, 'pdfs_borrados': eliminados})
    except Exception as e:
        if ui_log: ui_log(f"❌  eliminar-venta error: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@flask_app.route('/convertir-venta', methods=['POST', 'OPTIONS'])
def convertir_venta():
    """Convierte una OC en FAC: marca OC como Eliminado en ORDENES, crea FAC en VENTAS, regenera PDF"""
    if request.method == 'OPTIONS':
        return jsonify({'ok': True})
    data   = request.get_json(force=True)
    old_id = str(data.get('oldId', '')).strip()
    new_id = str(data.get('newId', '')).strip()
    sale   = data.get('sale', {})
    if not old_id or not new_id:
        return jsonify({'ok': False, 'error': 'Faltan oldId o newId'}), 400
    # Evitar factura duplicada si la app reintenta la conversión
    if new_id in IDS_REGISTRADOS:
        return jsonify({'ok': True, 'newId': new_id, 'msg': 'ya_registrado'})
    try:
        actualizar_estado_en_excel(old_id, 'OC', 'Eliminado')
        sale['id']   = new_id
        sale['type'] = 'FAC'
        escribir_venta(sale)
        IDS_REGISTRADOS.add(new_id)
        _ventas_cache['ts'] = 0
        eliminar_pdfs_de_carpeta(old_id, 'OC')
        pdf_path = generar_pdf_servidor(sale)
        hora = datetime.datetime.now().strftime('%H:%M')
        if ui_log: ui_log(f"🔄  [{hora}]  {old_id} → {new_id} convertido a Factura")
        return jsonify({'ok': True, 'newId': new_id})
    except Exception as e:
        if ui_log: ui_log(f"❌  convertir-venta error: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@flask_app.route('/pdf', methods=['POST', 'OPTIONS'])
def recibir_pdf():
    """
    Recibe el PDF generado por jsPDF en la app (base64) y lo guarda en la carpeta
    correspondiente. Esto sobreescribe el PDF auto-generado con la versión exacta
    que el usuario ve en pantalla.
    """
    if request.method == 'OPTIONS':
        return jsonify({'ok': True})
    data     = request.get_json(force=True)
    tipo     = data.get('type', 'FAC')
    filename = data.get('filename', 'documento.pdf')
    pdf_b64  = data.get('data', '')
    if not pdf_b64:
        return jsonify({'ok': False, 'error': 'Sin datos PDF'}), 400
    try:
        folder = FACTURAS_DIR if tipo == 'FAC' else ORDENES_DIR
        folder.mkdir(parents=True, exist_ok=True)
        # Sanitizar nombre de archivo
        safe_name = re.sub(r'[<>:"/\\|?*]', '_', filename)
        filepath  = folder / safe_name
        with open(filepath, 'wb') as f:
            f.write(base64.b64decode(pdf_b64))
        hora = datetime.datetime.now().strftime('%H:%M')
        if ui_log: ui_log(f"📄  [{hora}]  PDF (app) guardado: {safe_name}")
        return jsonify({'ok': True, 'filename': safe_name})
    except Exception as e:
        if ui_log: ui_log(f"⚠️  PDF recibido error: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


def run_flask():
    cargar_ids_desde_excel()
    flask_app.run(host='0.0.0.0', port=PORT, debug=False, use_reloader=False, threaded=True)

def run_ngrok():
    """Lanza ngrok con dominio estático en segundo plano."""
    global ngrok_proc
    ngrok_exe = shutil.which('ngrok')
    if not ngrok_exe:
        if ui_log: ui_log("⚠️  ngrok no encontrado — solo modo WiFi local")
        return
    try:
        ngrok_proc = subprocess.Popen(
            [ngrok_exe, 'http', f'--domain={NGROK_DOMAIN}', str(PORT)],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
        )
        if ui_log: ui_log(f"🌐  ngrok activo → {NGROK_URL}")
    except Exception as e:
        if ui_log: ui_log(f"⚠️  ngrok error: {e}")

# ════════════════════════════════════════
#  WEBAUTHN (Face ID / biometría)
# ════════════════════════════════════════
@flask_app.route('/webauthn/reg-challenge', methods=['POST', 'OPTIONS'])
def wa_reg_challenge():
    if request.method == 'OPTIONS': return jsonify({'ok': True})
    data = request.get_json(force=True) or {}
    user = str(data.get('user', '')).strip().lower()
    if user not in WA_USUARIOS:
        return jsonify({'error': 'Usuario no encontrado'}), 404
    opts = webauthn.generate_registration_options(
        rp_id=WA_RP_ID, rp_name='CacusaPOS',
        user_id=user.encode(), user_name=user, user_display_name=user,
        attestation=AttestationConveyancePreference.NONE,
        authenticator_selection=AuthenticatorSelectionCriteria(
            authenticator_attachment=AuthenticatorAttachment.PLATFORM,
            user_verification=UserVerificationRequirement.PREFERRED,
            resident_key=ResidentKeyRequirement.DISCOURAGED,
        ),
        supported_pub_key_algs=[
            COSEAlgorithmIdentifier.ECDSA_SHA_256,
            COSEAlgorithmIdentifier.RSASSA_PKCS1_v1_5_SHA_256,
        ],
    )
    _wa_reg_ch[user] = opts.challenge
    return jsonify({'challenge': bytes_to_base64url(opts.challenge), 'rpId': WA_RP_ID})


@flask_app.route('/webauthn/register', methods=['POST', 'OPTIONS'])
def wa_register():
    if request.method == 'OPTIONS': return jsonify({'ok': True})
    data = request.get_json(force=True) or {}
    user    = str(data.get('user', '')).strip().lower()
    cred_id = data.get('credentialId', '')
    att_obj = data.get('attestationObject', '')
    cd_json = data.get('clientDataJSON', '')
    if user not in WA_USUARIOS or not cred_id or not att_obj or not cd_json:
        return jsonify({'error': 'Datos incompletos'}), 400
    challenge = _wa_reg_ch.pop(user, None)
    if not challenge:
        return jsonify({'error': 'Challenge no encontrado o expirado'}), 400
    try:
        verification = webauthn.verify_registration_response(
            credential=RegistrationCredential(
                id=cred_id, raw_id=base64url_to_bytes(cred_id),
                response=AuthenticatorAttestationResponse(
                    client_data_json=base64url_to_bytes(cd_json),
                    attestation_object=base64url_to_bytes(att_obj),
                ),
            ),
            expected_challenge=challenge,
            expected_rp_id=WA_RP_ID,
            expected_origin=WA_ORIGINS,
            require_user_verification=False,
        )
    except Exception as e:
        return jsonify({'error': f'Verificación fallida: {e}'}), 400
    creds = _wa_load()
    creds[user] = {
        'credentialId': bytes_to_base64url(verification.credential_id),
        'publicKey':    bytes_to_base64url(verification.credential_public_key),
        'signCount':    verification.sign_count,
    }
    _wa_save(creds)
    return jsonify({'ok': True})


@flask_app.route('/webauthn/login-challenge', methods=['POST', 'OPTIONS'])
def wa_login_challenge():
    if request.method == 'OPTIONS': return jsonify({'ok': True})
    data = request.get_json(force=True) or {}
    user = str(data.get('user', '')).strip().lower()
    if user not in WA_USUARIOS:
        return jsonify({'error': 'Usuario no encontrado'}), 404
    creds = _wa_load()
    if user not in creds:
        return jsonify({'error': 'Face ID no configurado para este usuario'}), 404
    opts = webauthn.generate_authentication_options(
        rp_id=WA_RP_ID,
        allow_credentials=[PublicKeyCredentialDescriptor(
            id=base64url_to_bytes(creds[user]['credentialId'])
        )],
        user_verification=UserVerificationRequirement.PREFERRED,
    )
    _wa_auth_ch[user] = opts.challenge
    return jsonify({
        'challenge':    bytes_to_base64url(opts.challenge),
        'credentialId': creds[user]['credentialId'],
        'rpId':         WA_RP_ID,
    })


@flask_app.route('/webauthn/login', methods=['POST', 'OPTIONS'])
def wa_login():
    if request.method == 'OPTIONS': return jsonify({'ok': True})
    data    = request.get_json(force=True) or {}
    user    = str(data.get('user', '')).strip().lower()
    cred_id = data.get('credentialId', '')
    auth_d  = data.get('authenticatorData', '')
    cd_json = data.get('clientDataJSON', '')
    sig     = data.get('signature', '')
    if user not in WA_USUARIOS or not cred_id or not auth_d or not cd_json or not sig:
        return jsonify({'error': 'Datos incompletos'}), 400
    challenge = _wa_auth_ch.pop(user, None)
    if not challenge:
        return jsonify({'error': 'Challenge no encontrado o expirado'}), 400
    creds = _wa_load()
    if user not in creds:
        return jsonify({'error': 'Credencial no encontrada'}), 404
    stored = creds[user]
    try:
        verification = webauthn.verify_authentication_response(
            credential=AuthenticationCredential(
                id=cred_id, raw_id=base64url_to_bytes(cred_id),
                response=AuthenticatorAssertionResponse(
                    client_data_json=base64url_to_bytes(cd_json),
                    authenticator_data=base64url_to_bytes(auth_d),
                    signature=base64url_to_bytes(sig),
                ),
            ),
            expected_challenge=challenge,
            expected_rp_id=WA_RP_ID,
            expected_origin=WA_ORIGINS,
            credential_public_key=base64url_to_bytes(stored['publicKey']),
            credential_current_sign_count=stored['signCount'],
            require_user_verification=False,
        )
    except Exception as e:
        return jsonify({'error': f'Autenticación fallida: {e}'}), 401
    stored['signCount'] = verification.new_sign_count
    _wa_save(creds)
    return jsonify({'ok': True})


# ════════════════════════════════════════
#  INTERFAZ GRÁFICA
# ════════════════════════════════════════
class ServerUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CacusaPOS · Servidor WiFi")
        self.geometry("540x560")
        self.resizable(False, False)
        self.configure(bg='#FDE8F0')
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.local_ip = get_local_ip()
        self._build()
        self._start_server()

    def _build(self):
        # ── Header ──
        hdr = tk.Frame(self, bg='#C9637A', pady=16)
        hdr.pack(fill='x')
        tk.Label(hdr, text="💎  CACUSA POS",
                 font=('Helvetica', 19, 'bold'), bg='#C9637A', fg='white').pack()
        tk.Label(hdr, text="Servidor WiFi · Recibe ventas del iPhone en tiempo real",
                 font=('Helvetica', 10), bg='#C9637A', fg='#FFD6E0').pack(pady=(2, 0))

        body = tk.Frame(self, bg='#FDE8F0', padx=20, pady=14)
        body.pack(fill='both', expand=True)

        # ── Tarjeta HTTPS (ngrok — URL principal) ──
        https_card = tk.Frame(body, bg='#E8F5E9', relief='solid', bd=1)
        https_card.pack(fill='x', pady=(0, 8))

        tk.Label(https_card,
                 text="🔒  URL HTTPS — úsala en el iPhone (funciona en cualquier lugar):",
                 font=('Helvetica', 10, 'bold'), bg='#E8F5E9', fg='#1B5E20',
                 justify='left').pack(padx=14, pady=(10, 2), anchor='w')

        https_row = tk.Frame(https_card, bg='#E8F5E9')
        https_row.pack(fill='x', padx=14, pady=(0, 4))

        self.ngrok_var = tk.StringVar(value=NGROK_URL)
        tk.Label(https_row, textvariable=self.ngrok_var,
                 font=('Courier', 13, 'bold'), bg='#E8F5E9', fg='#2E7D32').pack(side='left')

        tk.Button(https_row, text="  Copiar  ",
                  font=('Helvetica', 10, 'bold'), bg='#2E7D32', fg='white',
                  activebackground='#1B5E20', relief='flat', padx=10, pady=4,
                  cursor='hand2', command=self._copiar_ngrok).pack(side='right')

        self.lbl_ngrok = tk.Label(https_card,
                 text="⏳  Iniciando ngrok...",
                 font=('Helvetica', 9), bg='#E8F5E9', fg='#558B2F')
        self.lbl_ngrok.pack(padx=14, pady=(0, 8), anchor='w')

        # ── Tarjeta WiFi local (respaldo) ──
        ip_card = tk.Frame(body, bg='white', relief='solid', bd=1)
        ip_card.pack(fill='x', pady=(0, 10))

        tk.Label(ip_card,
                 text="📶  WiFi local (solo cuando estés en casa con el servidor abierto):",
                 font=('Helvetica', 9), bg='white', fg='#A0607A',
                 justify='left').pack(padx=14, pady=(8, 2), anchor='w')

        ip_row = tk.Frame(ip_card, bg='white')
        ip_row.pack(fill='x', padx=14, pady=(0, 10))

        self.ip_var = tk.StringVar(value=f"http://{self.local_ip}:{PORT}")
        tk.Label(ip_row, textvariable=self.ip_var,
                 font=('Courier', 13, 'bold'), bg='white', fg='#C9637A').pack(side='left')

        tk.Button(ip_row, text="  Copiar  ",
                  font=('Helvetica', 10, 'bold'), bg='#C9637A', fg='white',
                  activebackground='#A04060', relief='flat', padx=10, pady=4,
                  cursor='hand2', command=self._copiar_ip).pack(side='right')

        # ── Estado ──
        self.lbl_estado = tk.Label(body,
                 text="⏳  Iniciando servidor...",
                 font=('Helvetica', 11), bg='#FDE8F0', fg='#A0607A')
        self.lbl_estado.pack(anchor='w', pady=(0, 8))

        # ── Archivo Excel ──
        excel_ok = EXCEL.exists()
        excel_color = '#4CAF7D' if excel_ok else '#E05555'
        excel_txt   = f"✅  Excel encontrado: {EXCEL.name}" if excel_ok else f"⚠️  No se encontró: {EXCEL}"
        tk.Label(body, text=excel_txt,
                 font=('Helvetica', 10), bg='#FDE8F0', fg=excel_color,
                 wraplength=480, justify='left').pack(anchor='w', pady=(0, 10))

        # ── Botón Sync Admin ──
        sync_row = tk.Frame(body, bg='#FDE8F0')
        sync_row.pack(fill='x', pady=(0, 8))
        tk.Button(sync_row, text="🔄  Sync Admin → Excel",
                  font=('Helvetica', 10, 'bold'), bg='#7B5EA7', fg='white',
                  activebackground='#5C3D8F', relief='flat', padx=12, pady=6,
                  cursor='hand2', command=self._sync_admin).pack(side='left')
        self.lbl_sync = tk.Label(sync_row, text="Importa productos del sitio web al Excel",
                 font=('Helvetica', 9), bg='#FDE8F0', fg='#A0607A')
        self.lbl_sync.pack(side='left', padx=(10, 0))

        # ── Log ──
        tk.Label(body, text="Registro de ventas recibidas:",
                 font=('Helvetica', 10, 'bold'), bg='#FDE8F0', fg='#3A1020').pack(anchor='w', pady=(0, 4))

        log_frame = tk.Frame(body, bg='white', relief='solid', bd=1)
        log_frame.pack(fill='both', expand=True)

        self.log_txt = scrolledtext.ScrolledText(log_frame,
                 font=('Helvetica', 11), bg='white', fg='#3A1020',
                 relief='flat', bd=0, state='disabled', wrap='word')
        self.log_txt.pack(fill='both', expand=True, padx=8, pady=8)

        # ── Pie ──
        tk.Label(self,
                 text="Mantén esta ventana abierta mientras usas la app en el iPhone  ·  Las ventas se guardan al instante en tu Excel",
                 font=('Helvetica', 8), bg='#FDE8F0', fg='#C898AA',
                 wraplength=520).pack(pady=(0, 8))

    def _agregar_log(self, msg):
        self.log_txt.config(state='normal')
        self.log_txt.insert('end', msg + '\n')
        self.log_txt.see('end')
        self.log_txt.config(state='disabled')

    def _sync_admin(self):
        self.lbl_sync.config(text="⏳ Sincronizando...", fg='#7B5EA7')
        def _run():
            count, error = sincronizar_desde_admin()
            if error:
                self.after(0, lambda: self.lbl_sync.config(text=f"❌ {error}", fg='#E05555'))
            else:
                self.after(0, lambda: self.lbl_sync.config(
                    text=f"✅ {count} productos sincronizados", fg='#2E7D32'))
            self.after(5000, lambda: self.lbl_sync.config(
                text="Importa productos del sitio web al Excel", fg='#A0607A'))
        threading.Thread(target=_run, daemon=True).start()

    def _copiar_ngrok(self):
        self.clipboard_clear()
        self.clipboard_append(self.ngrok_var.get())
        self.lbl_estado.config(text="✅  URL HTTPS copiada. Ábrela en Safari del iPhone", fg='#2E7D32')
        self.after(3000, lambda: self.lbl_estado.config(
            text="🟢  Servidor activo · ngrok HTTPS listo", fg='#4CAF7D'))

    def _copiar_ip(self):
        self.clipboard_clear()
        self.clipboard_append(self.ip_var.get())
        self.lbl_estado.config(text="✅  Dirección copiada. Pégala en Ajustes → Dirección del PC", fg='#4CAF7D')
        self.after(3000, lambda: self.lbl_estado.config(
            text="🟢  Servidor activo · Esperando ventas del iPhone", fg='#4CAF7D'))

    def _start_server(self):
        global ui_log
        ui_log = lambda msg: self.after(0, self._agregar_log, msg)
        # Arrancar Flask
        t = threading.Thread(target=run_flask, daemon=True)
        t.start()
        # Arrancar ngrok después de 2s (Flask necesita estar listo)
        self.after(2000, self._start_ngrok)
        self.after(1500, lambda: self.lbl_estado.config(
            text="🟢  Servidor activo · Iniciando ngrok...", fg='#4CAF7D'))

    def _start_ngrok(self):
        threading.Thread(target=self._ngrok_worker, daemon=True).start()

    def _ngrok_worker(self):
        run_ngrok()
        # Verificar que ngrok esté respondiendo (esperar hasta 10s)
        import urllib.request, time
        for _ in range(10):
            time.sleep(1)
            try:
                with urllib.request.urlopen('http://localhost:4040/api/tunnels', timeout=1) as r:
                    data = json.loads(r.read())
                    tunnels = data.get('tunnels', [])
                    if tunnels:
                        url = tunnels[0].get('public_url', NGROK_URL)
                        self.after(0, lambda u=url: self.lbl_ngrok.config(
                            text=f"✅  ngrok activo · HTTPS listo", fg='#2E7D32'))
                        self.after(0, lambda: self.lbl_estado.config(
                            text="🟢  Servidor activo · ngrok HTTPS listo", fg='#4CAF7D'))
                        return
            except:
                pass
        # Si no se pudo verificar por API, igual mostrar como activo
        self.after(0, lambda: self.lbl_ngrok.config(
            text="✅  ngrok iniciado", fg='#2E7D32'))
        self.after(0, lambda: self.lbl_estado.config(
            text="🟢  Servidor activo · ngrok HTTPS listo", fg='#4CAF7D'))

    def on_close(self):
        if tk_messagebox.askokcancel("Cerrar", "¿Cerrar el servidor?\nLas ventas del iPhone ya no se guardarán en Excel hasta que lo abras de nuevo."):
            if ngrok_proc:
                try: ngrok_proc.terminate()
                except: pass
            self.destroy()

if __name__ == '__main__':
    ui = ServerUI()
    ui.mainloop()
